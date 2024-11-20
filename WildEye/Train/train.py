# from ultralytics import YOLO

# # Load a model
# model = YOLO("yolov8n.yaml")  # build a new model from scratch
# model = YOLO("yolov8n.pt")  # load a pretrained model (recommended for training)

# # Use the model
# model.train(data="coco128.yaml", epochs=3)  # train the model
# metrics = model.val()  # evaluate model performance on the validation set
# results = model("https://ultralytics.com/images/bus.jpg")  # predict on an image
# path = model.export(format="tfjs")  # export the model to TenserFlowJS format





# import cv2
# import torch
# import numpy as np
# import pandas as pd
# from datetime import datetime
# import os
# from openpyxl import load_workbook
# from openpyxl.utils.dataframe import dataframe_to_rows

# # Load your YOLOv5 custom model
# model = torch.hub.load('ultralytics/yolov5', 'custom', path='/Users/apoorvakhajbage/Desktop/Apoorva/college/BTECH/ADS+CV/wild-eye/WildEye/Train/best.pt', force_reload=True)

# # Initialize video capture
# video_path = "/Users/apoorvakhajbage/Desktop/Apoorva/college/BTECH/ADS+CV/wild-eye/WildEye/Frontend/public/demo/monkey.mp4"
# cap = cv2.VideoCapture(video_path)

# # Initialize a dictionary to count animals
# animal_counts = {}
# previous_detections = []

# # Excel file path
# excel_file = "animal_counts.xlsx"

# # Function to update the Excel file with filtered animal counts
# def update_excel(animal_counts, max_animals, other_counts):
#     data = []
#     timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#     # Add max count animals
#     for animal in max_animals:
#         data.append([animal, animal_counts[animal], timestamp])

#     # Add animals exceeding threshold (>3)
#     for animal, count in other_counts.items():
#         data.append([animal, count, timestamp])

#     # Convert to DataFrame
#     df = pd.DataFrame(data, columns=["Animal Name", "Count", "Timestamp"])

#     if os.path.exists(excel_file):
#         # Load the existing Excel file
#         book = load_workbook(excel_file)
#         sheet = book.active

#         # Append new data below the last row
#         for r in dataframe_to_rows(df, index=False, header=False):
#             sheet.append(r)

#         # Save the updated Excel file
#         book.save(excel_file)

#     else:
#         # If the file doesn't exist, create a new file with headers
#         df.to_excel(excel_file, index=False)

# def calculate_iou(box1, box2):
#     """Calculate the Intersection over Union (IoU) of two bounding boxes."""
#     x1_inter = max(box1[0], box2[0])
#     y1_inter = max(box1[1], box2[1])
#     x2_inter = min(box1[2], box2[2])
#     y2_inter = min(box1[3], box2[3])

#     inter_area = max(0, x2_inter - x1_inter) * max(0, y2_inter - y1_inter)
#     box1_area = (box1[2] - box1[0]) * (box1[3] - box1[1])
#     box2_area = (box2[2] - box2[0]) * (box2[3] - box2[1])

#     union_area = box1_area + box2_area - inter_area
#     return inter_area / union_area if union_area > 0 else 0

# while cap.isOpened():
#     ret, frame = cap.read()
#     if not ret:
#         break

#     # Perform detection
#     results = model(frame)
#     detections = results.xyxy[0]  # Get predictions for the current frame

#     # Prepare to track detections
#     current_detections = []

#     for *bbox, conf, cls in detections:
#         if conf < 0.5:  # Filter by confidence threshold
#             continue
        
#         class_id = int(cls.item())
#         class_name = model.names[class_id]

#         # Format detection: [x1, y1, x2, y2, class_name]
#         current_detections.append((*bbox, class_name))

#         # Draw bounding boxes on the frame
#         x1, y1, x2, y2 = map(int, bbox)
#         cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
#         cv2.putText(frame, f"{class_name} {conf:.2f}", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

#     # Count unique animals using IoU
#     for det in current_detections:
#         x1, y1, x2, y2, class_name = det
#         new_box = (x1, y1, x2, y2)
        
#         # Check for overlaps with previous detections
#         is_new = True
#         for prev_det in previous_detections:
#             prev_box, prev_class_name = prev_det
#             if prev_class_name == class_name and calculate_iou(new_box, prev_box) > 0.5:
#                 is_new = False
#                 break

#         if is_new:
#             # Count this animal
#             if class_name in animal_counts:
#                 animal_counts[class_name] += 1
#             else:
#                 animal_counts[class_name] = 1

#         # Add current detection to previous detections
#         previous_detections.append((new_box, class_name))

#     # Display the frame with bounding boxes
#     cv2.imshow('YOLOv5 Detection with Improved Tracking', frame)

#     # Break the loop if the 'q' key is pressed
#     if cv2.waitKey(1) & 0xFF == ord('q'):
#         break

# # Release video capture and close windows
# cap.release()
# cv2.destroyAllWindows()

# # Print only the maximum count and any counts greater than 3
# max_count = max(animal_counts.values(), default=0)
# max_animals = [animal for animal, count in animal_counts.items() if count == max_count]

# # Find any animals with counts greater than 3, excluding the max animal
# other_counts = {animal: count for animal, count in animal_counts.items() if count > 3 and animal not in max_animals}

# print("Animal Counts:")
# for animal in max_animals:
#     print(f"{animal}: {max_count}")

# for animal, count in other_counts.items():
#     print(f"{animal}: {count}")

# # Update the Excel sheet with only max animals and those exceeding the threshold
# update_excel(animal_counts, max_animals, other_counts)

import cv2
import torch
import numpy as np
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from cassandra.cluster import Cluster  # Import Cassandra driver

# Load your YOLOv5 custom model
model = torch.hub.load('ultralytics/yolov5', 'custom', path='/Users/apoorvakhajbage/Desktop/Apoorva/college/BTECH/ADS+CV/wild-eye/WildEye/Train/best.pt', force_reload=True)

# Initialize video capture
video_path = "/Users/apoorvakhajbage/Desktop/Apoorva/college/BTECH/ADS+CV/wild-eye/WildEye/Frontend/public/demo/elephant.mp4"
cap = cv2.VideoCapture(video_path)

# Initialize a dictionary to count animals
animal_counts = {}
previous_detections = []

# Excel file path
excel_file = "animal_counts.xlsx"

# Connect to Cassandra
cluster = Cluster(['127.0.0.1'])  # Replace with your Cassandra cluster IP
session = cluster.connect('test')  # Replace with your keyspace name

# Function to create table if it doesn't exist
def create_table():
    session.execute("""
    CREATE TABLE IF NOT EXISTS animal_counts (
        animal_name text,
        count int,
        timestamp text,
        PRIMARY KEY (animal_name, timestamp)
    )
    """)

# Call the create_table function
create_table()

# Function to update the Excel file with filtered animal counts
def update_excel(animal_counts, max_animals, other_counts):
    data = []
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Add max count animals
    for animal in max_animals:
        data.append([animal, animal_counts[animal], timestamp])

    # Add animals exceeding threshold (>3)
    for animal, count in other_counts.items():
        data.append([animal, count, timestamp])

    # Convert to DataFrame
    df = pd.DataFrame(data, columns=["Animal Name", "Count", "Timestamp"])

    if os.path.exists(excel_file):
        # Load the existing Excel file
        book = load_workbook(excel_file)
        sheet = book.active

        # Append new data below the last row
        for r in dataframe_to_rows(df, index=False, header=False):
            sheet.append(r)

        # Save the updated Excel file
        book.save(excel_file)

    else:
        # If the file doesn't exist, create a new file with headers
        df.to_excel(excel_file, index=False)

    # Insert data into Cassandra
    for row in data:
        session.execute(
            """
            INSERT INTO animal_counts (animal_name, count, timestamp)
            VALUES (%s, %s, %s)
            """, (row[0], row[1], row[2])
        )

def calculate_iou(box1, box2):
    """Calculate the Intersection over Union (IoU) of two bounding boxes."""
    x1_inter = max(box1[0], box2[0])
    y1_inter = max(box1[1], box2[1])
    x2_inter = min(box1[2], box2[2])
    y2_inter = min(box1[3], box2[3])

    inter_area = max(0, x2_inter - x1_inter) * max(0, y2_inter - y1_inter)
    box1_area = (box1[2] - box1[0]) * (box1[3] - box1[1])
    box2_area = (box2[2] - box2[0]) * (box2[3] - box2[1])

    union_area = box1_area + box2_area - inter_area
    return inter_area / union_area if union_area > 0 else 0

while cap.isOpened():
    ret, frame = cap.read()
    if not ret:
        break

    # Perform detection
    results = model(frame)
    detections = results.xyxy[0]  # Get predictions for the current frame

    # Prepare to track detections
    current_detections = []

    for *bbox, conf, cls in detections:
        if conf < 0.5:  # Filter by confidence threshold
            continue
        
        class_id = int(cls.item())
        class_name = model.names[class_id]

        # Format detection: [x1, y1, x2, y2, class_name]
        current_detections.append((*bbox, class_name))

        # Draw bounding boxes on the frame
        x1, y1, x2, y2 = map(int, bbox)
        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv2.putText(frame, f"{class_name} {conf:.2f}", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

    # Count unique animals using IoU
    for det in current_detections:
        x1, y1, x2, y2, class_name = det
        new_box = (x1, y1, x2, y2)
        
        # Check for overlaps with previous detections
        is_new = True
        for prev_det in previous_detections:
            prev_box, prev_class_name = prev_det
            if prev_class_name == class_name and calculate_iou(new_box, prev_box) > 0.5:
                is_new = False
                break

        if is_new:
            # Count this animal
            if class_name in animal_counts:
                animal_counts[class_name] += 1
            else:
                animal_counts[class_name] = 1

        # Add current detection to previous detections
        previous_detections.append((new_box, class_name))

    # Display the frame with bounding boxes
    cv2.imshow('YOLOv5 Detection with Improved Tracking', frame)

    # Break the loop if the 'q' key is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release video capture and close windows
cap.release()
cv2.destroyAllWindows()

# Print only the maximum count and any counts greater than 3
max_count = max(animal_counts.values(), default=0)
max_animals = [animal for animal, count in animal_counts.items() if count == max_count]

# Find any animals with counts greater than 3, excluding the max animal
other_counts = {animal: count for animal, count in animal_counts.items() if count > 3 and animal not in max_animals}

print("Animal Counts:")
for animal in max_animals:
    print(f"{animal}: {max_count}")

for animal, count in other_counts.items():
    print(f"{animal}: {count}")

# Update the Excel sheet with only max animals and those exceeding the threshold
update_excel(animal_counts, max_animals, other_counts)

# Clean up Cassandra connection
session.shutdown()
cluster.shutdown()
